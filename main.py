import os
import json
from pathlib import Path
from openpyxl import load_workbook
import pandas as pd
import regex as re
import sys

arguments = os.sys.argv

# debug
debug = False
if ('-d' in arguments) or ('--debug' in arguments):
    debug = True

if not os.path.exists("./output"):
    print("Creating output directory...")
    os.makedirs("./output")

## Find Query String
# read all files in descriptions
files = os.listdir("descriptions")
# get file containing 'ARCTIC_TERMS*'
# file = [f for f in files if f.startswith("ARCTIC_TERMS")][0]
file = [f for f in files if f.startswith("TERMS")][0]

# read xlsx file no header
print("Reading file:", file)
# wb = load_workbook(filename="descriptions/" + file, read_only=True)
# load_workbook with column names
wb = load_workbook(filename="descriptions/" + file, read_only=True)

# sheet = wb.worksheets[3]
sheet = wb.worksheets[2]
# convert to dataframe
# df = pd.DataFrame(sheet.values)
# convert to dataframe with column names from first row
df = pd.DataFrame(sheet.values, columns=next(sheet.values))[1:]

# remove empty rows
# df = df.dropna(how='all')


print(df.dropna(how='all'))
strings = {}
# RAW_QUERY = df.iloc[14,1]
# RAW_QUERY = df.iloc[11,1]
# RAW_QUERY = df.iloc[9,1]
# RAW_QUERY = df.iloc[8,1]
# RAW_QUERY = df.iloc[3,1] # PROBLEM TODO
# print(df.iloc[3,0], RAW_QUERY[0:30])
RAW_QUERY = df.iloc[4,1] # PROBLEM TODO
print(df.iloc[4,0], RAW_QUERY[0:30])
# RAW_QUERY = df.iloc[6,1] # PROBLEM TODO
# print(df.iloc[6,0], RAW_QUERY[0:30])
# RAW_QUERY = df.iloc[2,1]
# RAW_QUERY = df.iloc[1,1]
# RAW_QUERY = df.iloc[0,1]

# RAW_QUERY = df.iloc[14,1]
print(RAW_QUERY[0:100])
input("Press Enter to continue...")
# RAW_QUERY = df.iloc[14,1]
# for i in range(len(df)):
#     RAW_QUERY = df.iloc[i,0]
#     if not RAW_QUERY or not isinstance(RAW_QUERY, str):
#         continue
#     if df.iloc[i,3] == 'Merge' or df.iloc[i,3] == 'Final':    
#         # print(RAW_QUERY[0:100])
#         # for keys in strings.keys():
#         # reverse order
#         for key in sorted(strings.keys(), key=len, reverse=True):
#             # print(key)
#             RAW_QUERY = RAW_QUERY.replace(key, strings[key])
#         # print(RAW_QUERY[0:100])
#     key = df.iloc[i,1]
#     if key == None or not isinstance(key, str):
#         continue
#     strings[key] = RAW_QUERY


if debug:
    RAW_QUERY = '( TITLE-ABS ( arctic  OR  tundra OR "PARANT"   )  AND  ( PUBYEAR  >  2004  AND  PUBYEAR  <  2020   )  AND NOT  DOCTYPE ( er ) )  OR  ( AUTHKEY ( arctic  OR  tundra   )  AND  ( PUBYEAR  >  2004  AND  PUBYEAR  <  2020   )  AND NOT  DOCTYPE ( er   )   )'        

    # RAW_QUERY = 'TITLE-ABS-KEY("machine learning" AND "healthcare") AND PUBYEAR > 2020' 
from src.elsapy import Scopus
scopus = Scopus(RAW_QUERY)
# scopus.search()

input("Press Enter to continue to OpenAlex...")

from src.openAlex import OpenAlex
openalex = OpenAlex(RAW_QUERY, debug)
# openalex.rundown()
